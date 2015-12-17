import openpyxl

testFileName = 'TestLocalization.xlsx'
fileName = 'LocalizationDataForKidMode.xlsx'
wb = openpyxl.load_workbook(fileName)

names = wb.get_sheet_names()
print names

sheet = wb.active
print type(sheet)
print sheet.title

print 'total rows: ' + str(sheet.max_row)
print 'total columns: ' + str(sheet.max_column)

for i in range(1, sheet.max_column):
    print 'column' + str(i) + ': ' + sheet.cell(row=1, column = i).value
print '~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~'

errorFileName = 'ErrorKey.txt'
errorList = []

for col in range(2, sheet.max_column):
    print 'column' + str(col) + ': ' + sheet.cell(row=1, column = col).value
    language = sheet.cell(row=1, column=col).value
    txtName = 'localization' + language + '.txt'
    
    errorList.append(language)
    number = 1
    
    print 'writing ' + txtName
    file = open(txtName, "wb")
    file.write("{\n")
    for r in range(2, sheet.max_row+1):
        keyString = sheet.cell(row=r, column=1).value
        valueString = sheet.cell(row=r, column=col).value
        print 'key = ' + keyString
        if valueString == None:
            print 'value = none'
        else:
            print 'value = ' + valueString
        sub = '"'
        times = 0
        if valueString != None:
            times = valueString.count(sub)
        if times > 0:
            isFound = True
            beginString = 0
            errorFound = 0
            while True:
                
                endString = len(valueString)
#                print 'valueString begin:' +  str(beginString) + ' end: ' + str(endString)
#                print 'sub valueString: ' + valueString[beginString:endString]
                result = valueString.find(sub, beginString, endString)
#                print 'valueString found " = ' + str(result)
                if result == -1:
                    beginString = 0
                    if errorFound == 1:
#                        print 'error'
                        s = str(number) + "." + keyString
#                        print s
                        errorList.append(s)
                        number = number + 1
                    break
                else:
                    sub2 = "\\"
#                    if result == 0:
#                        print 'substring: none'
#                    else:
#                        print 'substring: ' + valueString[result-1:result]
                    result2 = valueString.find(sub2, result-1, result)
#                    print 'valueString: is found \ before " = ' + str(result2)
                    if result2 == -1:
                        errorFound = 1
#                        print 'errorFound: ' + str(errorFound)
                        valueString = valueString[:result] + "\\" + valueString[result:]
#                        print 'new valueString: ' + valueString
                        beginString = result + 2
                    else:
                        beginString = result + 1
#                print "\n"

#        print 'row' + str(r) + ': ' + keyString
        file.write("\"" + keyString.encode('utf8') + "\"")
        file.write(" : ")
        if valueString == None:
            file.write("\"\",")
        else:
#            print "~" + valueString + "~"
            file.write("\"" + valueString.encode('utf8') + "\",")
        file.write("\n")
    file.write("}\n")
    file.close()


errorListFile = open(errorFileName, "wb")
for item in errorList:
    print item
    errorListFile.write(item + "\n")
errorListFile.close()



