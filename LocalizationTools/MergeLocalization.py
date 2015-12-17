#!/usr/bin/python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import codecs

list = ["Key", "En", "Spanish", "TraChinese", "French", "German", "Italian", "Japanese", "KR"]

#["Key", "En", "Spanish", "TraChinese", "French", "German", "Italian", "Japanese", "KR"]

wb = Workbook()
dest_filename = 'localization.xlsx'
ws1 = wb.active
for number in range(1, len(list)+1):
    ws1.cell(row=1, column=number, value=list[number-1])

keyList = []

for item in list:
    if item != "Key":
        num = 1
        columnIndex = list.index(item)+1
        print columnIndex
        filename = 'localization' + item + '.txt'
        path = 'OriginalTextForMerge/' + filename
#        file = open(path, 'r')

        if item == "German":
            encoding = 'latin-1'
        else:
            encoding = 'utf-8'

        file = codecs.open(path, 'r', encoding=encoding)
    
        for line in file:
            newline = line.rstrip('\n')
            if newline != '{' and newline != '' and newline != ' ' and newline != '}':
                n = newline.find(':')
                key = newline[0:n]
                key = key.rstrip()
                key = key.rstrip('"')
                key = key.lstrip('"')
                
                value = newline[n+1:len(newline)]
                value = value.rstrip()
                if value.endswith('",'):
                    value = value[:-2]
                elif value.endswith('"'):
                    value = value[:-1]
                else:
                    print 'not remove ", or " in ' + value
                value = value.lstrip()
                value = value.lstrip('"')
#                if key == "German":
#                print str(num) + ":" + key + "~" + value

                if item == "En":
                    keyList.append(key)
                    ws1.cell(row=num+1, column=1, value=key)
                    ws1.cell(row=num+1, column=columnIndex, value=value)
                else:
                    isfound = 0
                    try:
                        number = keyList.index(key)
                        isfound = 1
                    except ValueError:
                        isfound = 0
                    if isfound == 1:
                        ws1.cell(row=number+2, column=columnIndex, value=value)
                    else:
                        keyList.append(key)
                        ws1.cell(row=len(keyList)+1, column=1, value=key)
                        ws1.cell(row=len(keyList)+1, column=columnIndex, value=value)
            
                num = num + 1

wb.save(filename = dest_filename)
